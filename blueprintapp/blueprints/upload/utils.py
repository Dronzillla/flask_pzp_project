from io import BytesIO
from openpyxl import load_workbook, Workbook
from werkzeug.datastructures import FileStorage
from typing import Optional


def is_valid_excel_file(file: FileStorage) -> Optional[Workbook]:
    """Checks if given file is valid format.

    Args:
        file (FileStorage): form.file.data variable from Flask-WTF form.

    Returns:
        bool: 'True' if given file is valid format file, 'False' if given file is not valid format file.
    """
    # Check if filname endswith .xlsm or .xlsx
    if not file.filename.endswith(".xlsm") or not not file.filename.endswith(".xlsx"):
        return False
    # Read the file in memory
    in_memory_file = BytesIO(file.read())
    try:
        workbook = load_workbook(
            filename=in_memory_file, keep_vba=False, data_only=True
        )
    except:
        return False
    # Validate excel file is valid based on the provided list of worksheet names
    sheet_names = [
        "Pradžia",
        "Poveikis VF",
        "Rezultatai",
        "Jautrumo analizė",
        "Scenarijų analizė",
        "Grafikas",
    ]
    # Get the list of sheet names in the workbook
    sheets_in_workbook = workbook.sheetnames
    # Check for the presence of each specified sheet name
    for sheet_name in sheet_names:
        if sheet_name not in sheets_in_workbook:
            return False
    return workbook
